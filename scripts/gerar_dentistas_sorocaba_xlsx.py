# -*- coding: utf-8 -*-
"""
Gera planilha Excel com a quantidade APROXIMADA de dentistas cadastrados por
bairro em Sorocaba/SP.

Fonte: dentistaspertodemim.com (guia de telefones/endereços), coletado em
jun/2026. Cada bairro corresponde à contagem "【N empresas】" exibida na página
do próprio bairro nessa fonte. É uma ESTIMATIVA: cobre os bairros que possuem
página dedicada na fonte (~243 de 362 estabelecimentos listados na cidade);
bairros menores, sem página própria, não entram. Não é um censo oficial.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# (bairro, dentistas aprox.) — coletado de dentistaspertodemim.com em jun/2026
DADOS = [
    ("Centro", 88),
    ("Parque Campolim", 36),
    ("Jardim Vergueiro", 25),
    ("Jardim Santa Rosália", 16),
    ("Vila Trujillo", 13),
    ("Vila Hortência", 9),
    ("Éden", 8),
    ("Jardim Paulistano", 8),
    ("Além Ponte", 6),
    ("Jardim Simus", 5),
    ("Vila Jardini", 5),
    ("Jardim Faculdade", 4),
    ("Vila Haro", 4),
    ("Wanel Ville", 4),
    ("Jardim América", 4),
    ("Vila Barcelona", 2),
    ("Jardim Maria do Carmo", 2),
    ("Aparecidinha", 1),
    ("Brigadeiro Tobias", 1),
    ("Jardim Europa", 1),
    ("Jardim Saira", 1),
]

TOTAL_CIDADE = 362  # total de estabelecimentos listados para Sorocaba na fonte

# Ordena por quantidade (desc)
dados = sorted(DADOS, key=lambda x: x[1], reverse=True)
soma_amostra = sum(q for _, q in dados)

AZUL = "0A2A66"
AZUL_CLARO = "DCE6F5"
VERDE = "34C759"
CINZA = "F2F2F7"

wb = Workbook()
ws = wb.active
ws.title = "Dentistas por bairro"

thin = Side(style="thin", color="C8C8C8")
borda = Border(left=thin, right=thin, top=thin, bottom=thin)

# Título
ws.merge_cells("A1:C1")
c = ws["A1"]
c.value = "Dentistas por bairro — Sorocaba/SP (estimativa)"
c.font = Font(bold=True, size=14, color=AZUL)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 24

# Subtítulo / fonte
ws.merge_cells("A2:C2")
c = ws["A2"]
c.value = ("Fonte: dentistaspertodemim.com (jun/2026). Estimativa por bairro — "
           "não é censo oficial. Cobre bairros com página própria na fonte.")
c.font = Font(size=9, italic=True, color="6B7280")
ws.row_dimensions[2].height = 16

# Cabeçalho da tabela (linha 4)
headers = ["Bairro", "Dentistas (aprox.)", "% da amostra"]
for j, h in enumerate(headers, start=1):
    cell = ws.cell(row=4, column=j, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=AZUL)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = borda
ws.row_dimensions[4].height = 20

# Linhas de dados
linha = 5
for i, (bairro, qtd) in enumerate(dados):
    pct = qtd / soma_amostra
    b1 = ws.cell(row=linha, column=1, value=bairro)
    b2 = ws.cell(row=linha, column=2, value=qtd)
    b3 = ws.cell(row=linha, column=3, value=pct)
    b2.alignment = Alignment(horizontal="center")
    b3.alignment = Alignment(horizontal="center")
    b3.number_format = "0.0%"
    if i % 2 == 0:
        for col in (1, 2, 3):
            ws.cell(row=linha, column=col).fill = PatternFill("solid", fgColor=CINZA)
    for col in (1, 2, 3):
        ws.cell(row=linha, column=col).border = borda
    linha += 1

# Total da amostra
t1 = ws.cell(row=linha, column=1, value=f"Total da amostra ({len(dados)} bairros)")
t2 = ws.cell(row=linha, column=2, value=soma_amostra)
t3 = ws.cell(row=linha, column=3, value=1.0)
for col, cell in zip((1, 2, 3), (t1, t2, t3)):
    cell.font = Font(bold=True, color=AZUL)
    cell.fill = PatternFill("solid", fgColor=AZUL_CLARO)
    cell.border = borda
    if col > 1:
        cell.alignment = Alignment(horizontal="center")
t3.number_format = "0.0%"
linha += 2

# Nota de cobertura
ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=3)
c = ws.cell(row=linha, column=1)
c.value = (f"Obs.: a fonte lista {TOTAL_CIDADE} estabelecimentos em Sorocaba no total. "
           f"Esta amostra soma {soma_amostra} ({soma_amostra*100//TOTAL_CIDADE}%) nos bairros com página própria; "
           f"os ~{TOTAL_CIDADE - soma_amostra} restantes estão espalhados por bairros menores não listados aqui.")
c.font = Font(size=9, italic=True, color="6B7280")
c.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[linha].height = 42

# Larguras
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 14
ws.freeze_panes = "A5"

saida = "Dentistas_Sorocaba_por_bairro.xlsx"
wb.save(saida)
print(f"OK: {saida} ({len(dados)} bairros, soma {soma_amostra})")
